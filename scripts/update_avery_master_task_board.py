from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(__file__).resolve().parents[1] / "Avery Industries - Master Task Board.xlsx"


def main():
    if not WORKBOOK.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK}")

    wb = load_workbook(WORKBOOK)
    today = datetime.now().strftime("%Y-%m-%d")
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ensure_daily_brief(wb, today, timestamp)
    append_executive_report(wb, today, timestamp)

    wb.save(WORKBOOK)
    print(f"Updated {WORKBOOK}")
    print(f"Daily Brief date: {today}")
    print("Executive Reports row appended.")


def ensure_daily_brief(wb, today, timestamp):
    ws = wb["Daily Brief"]
    rows = {
        "Report Date": today,
        "Last Updated": timestamp,
        "Today's Critical Tasks": '=FILTER(\'Agent Tasks\'!A:N,\'Agent Tasks\'!G:G="Critical")',
        "Pending Approvals": '=FILTER(Approvals!A:L,Approvals!I:I="Pending")',
        "Blocked Tasks": '=FILTER(\'Agent Tasks\'!A:N,\'Agent Tasks\'!H:H="Blocked")',
        "New Leads": '=FILTER(Leads!A:N,Leads!I:I="New")',
        "Working Tasks Count": '=COUNTIF(\'Agent Tasks\'!H:H,"Working")',
        "Pending Approval Count": '=COUNTIF(Approvals!I:I,"Pending")',
        "High Priority Task Count": '=COUNTIF(\'Agent Tasks\'!G:G,"High")',
        "Qualified Lead Count": '=COUNTIF(Leads!I:I,"Qualified")',
        "Recommended CEO Decisions": "Manual ATLAS recommendation. Limit to 1-3 decisions.",
    }

    existing = {}
    for row in range(2, ws.max_row + 1):
        section = ws.cell(row=row, column=1).value
        if section:
            existing[str(section)] = row

    for section, value in rows.items():
        row = existing.get(section)
        if row is None:
            row = ws.max_row + 1
            ws.cell(row=row, column=1).value = section
        ws.cell(row=row, column=2).value = value
        if not ws.cell(row=row, column=3).value:
            ws.cell(row=row, column=3).value = "Auto-refreshed by local ATLAS updater."


def append_executive_report(wb, today, timestamp):
    ws = wb["Executive Reports"]
    next_row = ws.max_row + 1
    report_id = f"RPT-{next_row - 1:04d}"
    values = [
        report_id,
        today,
        "Executive Operations",
        "ATLAS Prime",
        f"Daily board refresh completed at {timestamp}.",
        "Refreshed Daily Brief formulas and counters.",
        "Google Sheets live sync requires Drive connector create/write permission.",
        "Review Executive Queue, Approvals, blocked tasks, and new leads.",
        '=COUNTIF(Approvals!I:I,"Pending")',
        "Move one revenue task forward before expanding roadmap.",
    ]
    for col, value in enumerate(values, start=1):
        ws.cell(row=next_row, column=col).value = value


if __name__ == "__main__":
    main()
